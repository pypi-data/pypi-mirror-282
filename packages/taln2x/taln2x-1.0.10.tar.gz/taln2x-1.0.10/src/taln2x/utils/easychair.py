import os, sys, csv, html, re, json
from openpyxl    import load_workbook
from unicodedata import normalize
from tqdm        import tqdm
from langdetect  import detect
from io          import StringIO
from bs4         import BeautifulSoup
from nameparser  import HumanName
from datetime    import datetime
from .           import Event, Track, Article, Author, readpdf
      
def get_articles(papers, abstracts, indir, accept, track_obj, \
                 verbose, ignore_pdf, sessions, \
                 id_col, authors_col, title_col, idhal_col, file_col, \
                 accept_col, keywords_col):
    # This function processes articles of a given track.
    # The accept parameter refers to the text to be used for acceptance 
    # (default is ACCEPT).
    all_papers      = [] #list of strings (output of the function)
    articles        = '' #proper csv string computed from the input file
    title2id        = {} #mapping title   <-> paperid
    title2abstract  = get_abstracts(abstracts) if abstracts else {}

    # [preprocessing] First, let us read the list of submitted papers
    # To fix keywords' format (newlines) to produce a proper csv string
    with open(papers, 'r') as f:
        current_article = None
        for line in f:
            line = normalize('NFC',line.strip())
            #print(line)
            if line and line[0].isdigit():
                if current_article is not None: 
                    # we store the article before moving on
                    articles += current_article + '\n'
                # we are moving to another article
                current_article = line
                #print(current_article)
                line_content = line.split('\t')
                # at the same time, we update the title2id dictionary
                # normalization wrt spaces
                # see https://stackoverflow.com/questions/2077897/
                # substitute-multiple-whitespace-with-single-whitespace-in-python
                norm_title = ' '.join(line_content[2].strip().lower().split())
                title2id[norm_title] = line_content[0].strip()
            else: 
                # lines which do not start with a number are keywords to
                # be appended to a paper description
                if line and line[0] != '#': # to ignore comments
                    current_article+= line
        # last entry should not be forgotten (since 
        # accumulation is triggered by new entries):
        if current_article:
            articles += current_article 
        #print(articles)

    # [processing] Second, let us read the computed proper csv string
    # and retrieve the corresponding info (e.g. pdf for the English title)
    reader = csv.reader(StringIO(articles), delimiter='\t', quotechar='|')
    for row in tqdm(reader, total=len(articles.split('\n'))):
        #print(PAPERINFO)
        #print(row)
        article = Article()
        patterns = accept.split(';') # in case there are several acceptance keywords
        if any(list(map(lambda acc : re.match('.*'+acc.strip(), row[accept_col], re.IGNORECASE), patterns))):
            ## update session info
            if sessions and track_obj.sessions != '':
                art_session = row[accept_col].split('_')[1]  # format: ACCEPT_SessionID
                if art_session in track_obj.sessions.split(';'):
                    if hasattr(track_obj,art_session): # to get session's full name
                        article.__dict__['session'] = track_obj.__dict__[art_session] 
                    else:
                        print(f'[Warning] {track_obj.trackid} full name unset in event.yml', \
                              file=sys.stderr)
                    track_obj.update_sessions(art_session, row[id_col])
                else:
                    if verbose > 0:
                        print('[Warning] Unknown session: ' + \
                              art_session, file=sys.stderr)
            ## collect paper info
            article.__dict__['paperid'] = row[id_col]
            article.__dict__['authors'] = row[authors_col]
            article.__dict__['title']   = row[title_col]
            article.__dict__['idhal']   = row[idhal_col]
            article.__dict__['keywords']= row[keywords_col]
            title                       = article.__dict__['title']
            title_norm                  = ' '.join(title.lower().split())
            title_norm                  = normalize('NFC', title_norm)
            if abstracts and title_norm in title2abstract:
                article.__dict__['abstract']= title2abstract[title_norm]
            else:
                #print("**",title2abstract.keys())
                #print("***", row[title_col])
                if verbose > 1:
                    print(f"[Warning] abstract not found for paper {article.__dict__['paperid']}", file=sys.stderr)   
            # make sure the halid info actually is a valid hal identifier
            midhal=re.fullmatch('hal-[0-9]*',article.__dict__['idhal'])
            if midhal is None:
                article.__dict__['idhal'] = ''
            # if authors are missing in articles.csv, try to get them from the abstracts' file
            if article.__dict__['authors'] == '': 
                try: 
                    article.__dict__['authors'] = id2authors[article.__dict__['paperid']]
                except KeyError as ke:
                    if verbose > 0:
                        print('[Warning] no author found for article ' + \
                              str(article.__dict__['paperid']), file=sys.stderr)
            # get article filename 
            article.__dict__['url'] = os.path.join(os.getcwd(), indir, 'pdf', row[file_col].strip())
            #print(article.__dict__['url'])

            ## Detect the article's language from its title
            article.__dict__['language'] = detect(row[title_col])
            #print('***', article.__dict__['title'], article.__dict__['language'])

            ## Extract the article's secondary title from the pdf (if possible)
            pdf_path = os.path.join(indir, 'pdf', row[file_col].strip())
            if not os.path.exists(pdf_path):
                print('[Error] PDF file not found: ' + pdf_path, file=sys.stderr)
                print('Please check input files.', file=sys.stderr)
                continue 
            # Read info from pdf file (article, indir, path, default_title, ignore_pdf, verbose)
            readpdf.read_pdf(article, indir, pdf_path, row[file_col].strip(), ignore_pdf, verbose)
            article.__dict__['track'] = track_obj.volume
            all_papers.append(article)
    return all_papers


def get_abstracts(abstracts):
    # Function which collects abstracts of a given track
    title2abstract = {}
    # read the abstracts (beware it is an html file)    
    abstracts_type = os.path.splitext(abstracts)[1][1:] #do not forget to remove the dot in extension
    with open(abstracts,'r') as f:
        content = f.read()
        if abstracts_type == 'html':
            soup = BeautifulSoup(content, 'html.parser')
            #print(soup.prettify())
            for x in soup.find_all('div', attrs={'class':'paper'}):
                authors = x.find('span', attrs={'class': 'authors'}).get_text().strip()[:-1] 
                title   = x.find('span', attrs={'class': 'title'}).get_text().strip()
                abstract= x.findNext('div').get_text()[10:]
                title_norm = ' '.join(title.lower().split())
                title_norm = normalize('NFC', title_norm)
                title2abstract[title_norm] = abstract.replace('\n','')
        elif abstracts_type == 'json':
            with open(abstracts) as abstracts_file:
                papers = json.load(abstracts_file)
            for pid in papers:
                title      = papers[pid]['title']
                title_norm = ' '.join(title.lower().split())
                title_norm = normalize('NFC', title_norm)
                abstract   = papers[pid]['abstract']
                #print("**", title_norm)
                title2abstract[title_norm] = abstract
    return title2abstract


def get_authors(authors):
    # Function which collects the authors of a given track
    authors_type     = os.path.splitext(authors)[1][1:]
    all_authors      = {}
    if authors_type == 'xlsx':
        previous_paperid = None
        author_counter   = 1
        # First, let us read the input XLSX file to turn it into some csv
        wb = load_workbook(authors, read_only=True)
        sh = wb.get_sheet_by_name('All')
        for row in sh.iter_rows(min_row=2):
            #print('###',';'.join(list(map(lambda x: str(x.value), row))))
            if row[0].value is None:
                break
            values  = [cell.value for cell in row]  
            author  = Author()
            paperid = int(values[0])
            author.__dict__['paperid']     = str(paperid)
            if previous_paperid == paperid:
              author_counter += 1
            else:
                author_counter  = 1
            author.__dict__['rank']        = author_counter
            previous_paperid               = paperid
            if sh.max_column == 9: 
                #the excel file comes from a premium easychair account
                #the first and lastnames are in two columns
                fullname                       = HumanName(values[1] + ' ' + values[2])
                # "~" is used as an unsplittable space in first/last names
                author.__dict__['firstname']   = re.sub(r'([^\\])~', r'\1 ', fullname.first)
                author.__dict__['lastname']    = (fullname.middle + ' ' + fullname.last).strip()
                author.__dict__['email']       = values[3]
                author.__dict__['affiliation'] = html.unescape(values[5] + ', ' \
                                                               + values[4])              
            else:
                fullname                       = HumanName(values[1])
                # "~" is used as an unsplittable space in first/last names
                author.__dict__['firstname']   = re.sub(r'([^\\])~', r'\1 ', fullname.first)
                author.__dict__['lastname']    = (fullname.middle + ' ' + fullname.last).strip()
                author.__dict__['email']       = values[2]
                author.__dict__['affiliation'] = html.unescape(values[4] + ', ' \
                                                               + values[3])
            if paperid in all_authors.keys():
                all_authors[paperid].append(author)
            else:
                all_authors[paperid] = [author]
            #print('###', author)
    elif authors_type == 'json':
        with open(authors) as authors_file:
            papers = json.load(authors_file)
        for pid, pauthors in papers.items():
            for prank, paut in papers[pid].items():
                author = Author()             
                author.__dict__['paperid']     = str(pid)
                author.__dict__['rank']        = prank
                author.__dict__['firstname']   = papers[pid][prank]['firstname']
                author.__dict__['lastname']    = papers[pid][prank]['lastname']
                author.__dict__['email']       = papers[pid][prank]['email']
                author.__dict__['affiliation'] = papers[pid][prank]['affiliation']
                if int(pid) in all_authors.keys(): 
                    all_authors[int(pid)].append(author)
                else:
                    all_authors[int(pid)] = [author]
                #print('###', author)
    return all_authors


def get_data(e, indir, verbose, ignore_pdf, sessions, id_col, authors_col, title_col, \
             idhal_col, file_col, accept_col, keywords_col):
    for t in e.__dict__['tracks'].keys():
        print('Processing track:', t, file=sys.stderr)
        track_dir = os.path.join(indir, t)
        if not os.path.exists(track_dir):
            print(' [Warning] Directory missing for track ' + t + \
                  ' (track ignored).' , file=sys.stderr)
            #move on to the next track
        else: 
            papers_file   = os.path.join(track_dir, 'articles.csv')
            abstracts_file= None
            authors_file  = None
            if os.path.exists(os.path.join(track_dir, 'accepted.html')):
                abstracts_file= os.path.join(track_dir, 'accepted.html')
            elif os.path.exists(os.path.join(track_dir, 'abstracts.json')):
                abstracts_file= os.path.join(track_dir, 'abstracts.json')
            else:
                if verbose >= 0:
                    print('[Warning] neither accepted.html nor abstracts.json found for track ' + t + \
                          '\n(abstracts will be extracted from pdf articles and may be noisy)', \
                          file=sys.stderr)                
            if os.path.exists(os.path.join(track_dir, 'author_list.xlsx')):
                authors_file  = os.path.join(track_dir, 'author_list.xlsx')
            elif os.path.exists(os.path.join(track_dir, 'authors.json')):
                authors_file  = os.path.join(track_dir, 'authors.json')
            else:
                if verbose >= 0:
                    print('[Warning] neither author_list.xlsx nor authors.json found for track ' + t + \
                          '\n(authors will be extracted from articles.csv and be missing affiliations)', file=sys.stderr)                
            if not os.path.exists(papers_file):
                print('\n[Error] Missing articles.csv for track ' + t + \
                      '\n(track ignored).\n', \
                      file=sys.stderr)
                continue
            accept   = 'ACCEPT' #default case
            track_obj= e.__dict__['tracks'][t]
            #if accept flag is given in the event.yml file, use it            
            if track_obj.accept != '':
                accept = track_obj.accept
            articles = get_articles(papers_file, abstracts_file, \
                                    track_dir, accept, track_obj, \
                                    verbose, ignore_pdf, sessions, \
                                    id_col, authors_col, title_col, idhal_col, file_col, \
                                    accept_col, keywords_col)
            authors  = get_authors(authors_file) if authors_file else {}
            track_obj.articles = articles
            track_obj.authors  = authors
