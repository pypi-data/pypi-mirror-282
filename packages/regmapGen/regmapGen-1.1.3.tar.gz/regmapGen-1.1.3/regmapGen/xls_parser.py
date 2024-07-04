#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""XLS table parser
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .register_node import RegisterNode


class XLSParser(object):
    def get_hex(self, data):
        if isinstance(data, int):
            return "%X" % data
        try:
            int(data, 16)
            data = data.upper()
            return data[2:]
        except ValueError:
            print("Fatal: Invalid data format(%s), only decimal or hexadecimal with prefix '0x' supported" % (data))
        return None

    def get_sheet(self, name):
        self.sheet = load_workbook(name, data_only=True)['RegMap']

    def in_range(self, column, start, end):
        return ord(column) in range(ord(start), ord(end))

    def locate(self, start_column):
        s_row = [cell for cell in self.sheet[start_column] if cell.value == 'Block'][0].row
        s_reg = [cell for cell in self.sheet[s_row] if cell.value == 'Register'][0].column
        s_field = [cell for cell in self.sheet[s_row] if cell.value == 'Field'][0].column
        if isinstance(s_reg, int):
            s_reg = get_column_letter(s_reg)
        if isinstance(s_field, int):
            s_field = get_column_letter(s_field)
        return (s_row, s_reg, s_field)

    def get_header(self):
        self.header = {}
        self.header['block'] = 'A'
        self.start_row, self.header['register'], self.header['field'] = self.locate(self.header['block'])
        for cell in self.sheet[self.start_row]:
            if isinstance(cell.column, int):
                column = get_column_letter(cell.column)
            else:
                column = cell.column
            if cell.value == 'Offset':
                if self.in_range(column, self.header['block'], self.header['register']):
                    self.header['block_offset'] = column
                elif self.in_range(column, self.header['register'], self.header['field']):
                    self.header['reg_offset'] = column
                else:
                    print("Warning: Invalid 'Offset' column presented (Col: %s) %(column)")
            if cell.value == 'Width':
                self.header['width'] = column
            if cell.value == 'Access':
                if self.in_range(column, self.header['register'], self.header['field']):
                    self.header['reg_access'] = column
                elif self.in_range(column, self.header['field'], 'Z'):
                    self.header['field_access'] = column
                else:
                    print("Warning: Invalid 'Access' column presented (Col: %s) %(column)")
            if cell.value == 'Repeat':
                if self.in_range(column, self.header['block'], self.header['register']):
                    self.header['block_repeat'] = column
                elif self.in_range(column, self.header['register'], self.header['field']):
                    self.header['reg_repeat'] = column
                else:
                    print("Warning: Invalid 'Repeat' column presented (Col: %s) %(column)")
            if cell.value == 'HDL Path':
                self.header['hdl_path'] = column
            if cell.value == 'Description':
                if self.in_range(column, self.header['register'], self.header['field']):
                    self.header['reg_description'] = column
                elif self.in_range(column, self.header['field'], 'Z'):
                    self.header['field_description'] = column
                else:
                    print("Warning: Invalid 'Description' column presented (Col: %s) %(column)")
            if cell.value == 'Bits':
                self.header['bits'] = column
            if cell.value == 'Reset Value':
                self.header['reset'] = column
            if cell.value == 'Has Reset':
                self.header['has_reset'] = column
            if cell.value == 'Rand':
                self.header['rand'] = column
            if cell.value == 'Volatile':
                self.header['volatile'] = column
            if cell.value == 'Hardware':
                self.header['hardware'] = column

    def parse_bits(self, field, bits):
        try:
            start = int(bits)
            field.lsb_pos = start
            field.size = 1
        except ValueError:
            bit_n = [int(x) for x in bits.split(':')]
            start = min(bit_n)
            end = max(bit_n)
            field.lsb_pos = start
            field.size = end - start + 1

    def set_hdl(self, reg, field, value):
        field.hdl_path = value
        reg.has_hdl_path = True

    def set_attr(self, node, key, value):
        node.attrs[key] = value

    def flat_str(self, data):
        """Convert multiline strings into single line strings recursively."""
        if isinstance(data, dict):
            return {k: self.flat_str(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self.flat_str(v) for v in data]
        elif isinstance(data, str) and '\n' in data:
            return data.replace('\n', ' ')
        else:
            return data

    def parse_data(self):
        self.get_header()

        switch = {
            self.header['width']: lambda cell: self.set_attr(block, 'width', cell.value),
            self.header['block_offset']: lambda cell: self.set_attr(block, 'offset', self.get_hex(cell.value)),
            self.header['block_repeat']: lambda cell: self.set_attr(block, 'repeat', cell.value),
            self.header['reg_offset']: lambda cell: self.set_attr(reg, 'offset', self.get_hex(cell.value)),
            self.header['reg_access']: lambda cell: self.set_attr(reg, 'access', cell.value),
            self.header['reg_repeat']: lambda cell: self.set_attr(reg, 'repeat', cell.value),
            self.header['reg_description']: lambda cell: self.set_attr(reg, 'description', self.flat_str(cell.value)),
            self.header['bits']: lambda cell: self.parse_bits(field, cell.value),
            self.header['field_access']: lambda cell: self.set_attr(field, 'access', cell.value),
            self.header['reset']: lambda cell: self.set_attr(field, 'reset', self.get_hex(cell.value)),
            self.header['has_reset']: lambda cell: self.set_attr(field, 'has_reset', cell.value),
            self.header['rand']: lambda cell: self.set_attr(field, 'is_rand', cell.value),
            self.header['volatile']: lambda cell: self.set_attr(field, 'is_volatile', cell.value),
            self.header['hdl_path']: lambda cell: self.set_hdl(reg, field, cell.value),
            self.header['field_description']: lambda cell: self.set_attr(field, 'description', self.flat_str(cell.value)), # noqa E501
            self.header['hardware']: lambda cell: self.set_attr(field, 'hardware', cell.value),
        }

        rmap = {}

        for row in self.sheet[self.start_row + 1:self.sheet.max_row]:
            for cell in row:
                if isinstance(cell.column, int):
                    column = get_column_letter(cell.column)
                else:
                    column = cell.column
                if cell.value is None:
                    continue
                try:
                    switch[column](cell)
                except KeyError:
                    if column == self.header['block']:
                        block = RegisterNode(cell.value)
                        rmap[cell.value] = block
                        continue
                    if column == self.header['register']:
                        reg = RegisterNode(cell.value)
                        reg.has_hdl_path = False
                        reg.field_num = 0
                        block[cell.value] = reg
                        continue
                    if column == self.header['field']:
                        field = RegisterNode(cell.value)
                        reg.field_num += 1
                        field.index = reg.field_num - 1
                        reg[cell.value] = field
                        continue

        return rmap

    def build_res_field(self, name, lsb_pos, size):
        field = RegisterNode(name)
        field.lsb_pos = lsb_pos
        field.size = size
        field.access = 'RO'
        field.reset = 0
        field.has_reset = 1
        field.is_rand = 0
        field.is_volatile = 0
        field.reserved = True
        field.description = "Reserved bit(s)"
        return field

    def fill_reserved(self, data):
        """Fill the reserved fields in the register."""
        for block_name, block in data.items():
            for reg_name, reg in block.iter_items():
                bits = {}
                exclusive_field = False
                for field_name, field in reg.iter_items():
                    # exclusive fields, no more process
                    if field.size == block.attrs['width']:
                        field.attrs['reserved'] = False
                        exclusive_field = True
                        break
                    bits[field.lsb_pos] = field.size
                    field.attrs['reserved'] = False

                if not exclusive_field:
                    # ordered by field lsb position
                    bits_tup = sorted(bits.items(), key=lambda x: x[0])
                    res_num = 0
                    # if the position of field is not 0, add a reserved field
                    if bits_tup[0][0] != 0:
                        res_name = "res_0"
                        res_field = self.build_res_field(res_name, 0, bits_tup[0][0])
                        res_num = 1
                        reg[res_name] = res_field
                    for idx in range(len(bits_tup)):
                        # if it's the last field, the upper limit of field is set to register width
                        if idx + 1 == len(bits_tup):
                            upper = block.attrs['width']
                        # normal field, the upper limit set to position of the adjacent higher field
                        else:
                            upper = bits_tup[idx + 1][0]
                        # if the bit occupy of lower field overlaps the adjacent higher field, report error
                        if bits_tup[idx][0] + bits_tup[idx][1] > upper:
                            print("Error: Field address conflict in register %s" % reg_name)
                        # if there is a gap between lower field and the adjacent higher field, add a reserved field
                        elif bits_tup[idx][0] + bits_tup[idx][1] < upper:
                            res_name = "res_%0d" % res_num
                            res_num += 1
                            res_field = self.build_res_field(
                                res_name,
                                bits_tup[idx][0] + bits_tup[idx][1],
                                upper - (bits_tup[idx][0] + bits_tup[idx][1])
                            )
                            reg[res_name] = res_field
        return data

    def reorder_by_lsb(self, data):
        """Reorders the fields by LSB."""
        for block_name, block in data.items():
            for reg_name, reg in block.iter_items():
                order_fields = [RegisterNode("placeholder")] * block.attrs['width']
                name_fields = []
                for field_name, field in reg.iter_items():
                    order_fields[field.lsb_pos] = field
                    name_fields.append(field.name)
                for name in name_fields:
                    del reg[name]
                reg_reset = 0
                order_fields.reverse()
                for field in order_fields:
                    if field.name != "placeholder":
                        reg[field.name] = field
                        if isinstance(field.attrs['reset'], str):
                            reg_reset += int(field.attrs['reset'], 16) << field.lsb_pos
                        else:
                            reg_reset += field.attrs['reset'] << field.lsb_pos
                reg.attrs['reset'] = reg_reset
        return data


def is_node(item):
    return isinstance(item, RegisterNode)


def get_bit_reset(reset, pos):
    return (reset & (1 << pos)) >> pos


def format_hex(block_offset, reg_offset):
    if isinstance(block_offset, str):
        block_offset_hex = int(block_offset, 16)
    else:
        block_offset_hex = block_offset
    if isinstance(reg_offset, str):
        reg_offset_hex = int(reg_offset, 16)
    else:
        reg_offset_hex = reg_offset
    return "0x%08X" % (block_offset_hex + reg_offset_hex)
