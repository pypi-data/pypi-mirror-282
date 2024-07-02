import datetime
import decimal
import json
import os
import re
import sys
import pwinput
import openpyxl
import sqldrafter
from prompt_toolkit import prompt
from openpyxl import load_workbook
from sqldrafter import analysisField
from sqldrafter import analysisUniq
from sqldrafter import analysisPartion
from sqldrafter import analysisJoinInfo

USAGE_STRING = """
使用: sqldrafter <command>
可使用命令如下:
    init\t\t\t初始化数据源
    createSpace\t\t\t初始化空间信息返回空间ID
    gen <spaceID> <table1> <table2>\t 为使用的表生成execle
    update <file>\t\t 更新必要的表信息给sqldrater
    query <spaceID> <question>\t\t\t 进行一个查询
    indicator <spaceId> <key> <value> \t\t 更新一个指标定义
"""

home_dir = os.path.expanduser("~")


def main():
    if len(sys.argv) < 2:
        print(USAGE_STRING)
        sys.exit(1)
    if sys.argv[1] == "init":
        init()
    elif sys.argv[1] == "createSpace":
        createSpace()
    elif sys.argv[1] == "gen":
        gen()
    elif sys.argv[1] == "update":
        update_schema()
    elif sys.argv[1] == "indicator":
        update_indicator()
    elif sys.argv[1] == ("query"):
        query()
    else:
        print(f"不支持的命令: {sys.argv[1]}")
        print(USAGE_STRING)
        sys.exit(1)


def init():
    """
    初始化数据源配置在 ~/.sqldrafter/connection.json
    """
    print("Welcome to \033[94msqldrafter\033[0m!\n")
    filepath = os.path.join(home_dir, ".sqldrafter", "connection.json")
    if os.path.exists(filepath):
        print(
            "目前已有数据源配置，请问是否覆盖? (y/n)"
        )
        overwrite = prompt().strip()
        if overwrite.lower() != "y":
            print("目前不覆盖，配置无变化！")
            sys.exit(0)
        else:
            print("我们将创建新的数据源配置在 ~/.sqldrafter/connection.json")
    else:
        print("我们将创建新的数据源配置在 ~/.sqldrafter/connection.json")
        if not os.path.exists(os.path.join(home_dir, ".sqldrafter")):
            os.mkdir(os.path.join(home_dir, ".sqldrafter"))

    if os.environ.get("SQLDRAFTER_API_KEY"):
        print(
            "我们发现您的 SQLDRAFTER_API_KEY 存在您的环境中. 我们将使用它。"
        )
        api_key = os.environ.get("SQLDRAFTER_API_KEY")
    else:
        print(
            "请输入SQLDRAFTER_API_KEY. 您可以在 https://www.sqldrafter.com/person-center/api-key/ 获取"
        )
        api_key = prompt().strip()

    # prompt user for db_type
    print(
        "请选择一个数据库类型 "
        + ", ".join(sqldrafter.SUPPORTED_DB_TYPES)
    )
    db_type = prompt().strip()
    db_type = db_type.lower()
    while db_type not in sqldrafter.SUPPORTED_DB_TYPES:
        print(
            "您输入的数据库，我们并不支持，目前只支持如下数据库： "
            + ", ".join(sqldrafter.SUPPORTED_DB_TYPES)
        )
        db_type = prompt().strip()
        db_type = db_type.lower()
    print("请输入数据库地址:")
    host = prompt().strip()
    print("请输入数据库端口:")
    port = prompt().strip()
    print("请输入数据库用户名:")
    user = prompt().strip()
    print("请输入数据库密码:")
    password = pwinput.pwinput(prompt="Please enter your database password:")
    if db_type == "hive" or db_type == "spark" :
        db_creds = {
            "host": host,
            "port": port,
            "username": user,
            "password": password,
        }
    else:
        db_creds = {
            "host": host,
            "port": port,
            "user": user,
            "password": password,
        }

    sqldrafter.sqlDrafter(api_key=api_key, db_type=db_type, db_creds=db_creds)
    # write to filepath and print confirmation
    with open(filepath, "w") as f:
        data = {"api_key": api_key, "db_type": db_type, "db_creds": db_creds}
        json.dump(data, f, indent=4)
    print(f"数据源配置将保存在  {filepath}.")
    sys.exit(0)


def createSpace():
    """
    初始化数据源配置在 ~/.sqldrafter/connection.json
    """
    sq = sqldrafter.sqlDrafter()
    # print welcome message
    print("请输入空间标题：")
    title = prompt().strip()
    print("请输入空间描述：")
    desc = prompt().strip()
    spaceID = sq.create_space(title, desc, 2)
    if spaceID <= 0:
        pass
    else:
        print("该空间ID为：" + str(spaceID) + "请妥善保存。")
    sys.exit(0)


def gen():
    sq = sqldrafter.sqlDrafter()

    if len(sys.argv) < 3:
        print("请输入空间ID：")
        space_id = prompt().strip()
    else:
        space_id = sys.argv[2]
    if space_id.isdigit():
        pass
    else:
        print("空间ID须为数字")
        sys.exit(0)
    if len(sys.argv) < 4:
        print(
            "输入需要的表格"
        )
        table_names = prompt().strip()
        table_name_list = re.split(r"\s+", table_names.strip())
    else:
        table_name_list = sys.argv[3:]
    if table_name_list == [""] or table_name_list == []:
        print("没有表格，将退出！")
        sys.exit(0)

    sql = sq.generate_db_schema(table_name_list)
    request = sqldrafter.addMultiTableRequest(int(space_id), sql)
    is_add = sq.add_table(json.dumps(request, cls=JSONSelfEncoder))
    if is_add:
        detail = sq.get_detail(int(space_id))
        create_excel(detail)
        pass
    sys.exit(0)


def create_excel(detail: json):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    data = detail["data"]
    sheet1.title = "空间"
    content = ["标题：", data["name"]]
    sheet1.append(content)
    content = ["描述：", data["busiDesc"]]
    sheet1.append(content)
    content = ["ID", to_str(data["id"])]
    sheet1.append(content)
    create_table_sheet(data["tableList"], workbook)
    workbook.save(to_str(data["id"]) + "space.xlsx")


def create_table_sheet(table_data: json
                       , workbook: openpyxl.Workbook):
    for item in table_data:
        sheet = workbook.create_sheet(item["tableName"])
        sheet.append(["ID：", item["id"]])
        sheet.append(["数据库名",item["dbName"]])
        sheet.append(["表名", item["tableName"]])
        sheet.append(["中文名：", item["tableCnName"]])
        sheet.append(["业务描述", item["tableDesc"]])
        sheet.append(["表格详情"])
        sheet.append(["是否主键", "列名", "数据类型", "中文名", "取值样例", "其他备注", "是否使用(0/1)"])
        for field in item["fields"]:
            sheet.append(
                [to_str(is_None(field['primaryKey'])), to_str(is_None(field['name'])), to_str(is_None(field["type"])),
                 to_str(is_None(field["cnName"])),
                 to_str(is_None(field["samples"])), to_str(is_None(field["remarks"])), to_str(is_None(field["used"]))])
        create_uniquekey(item["uniqueKeyInfoList"], sheet)
        create_join_info(item["joinInfoList"], sheet)
        create_partition_info(item["partitionFieldInfoList"],sheet)

def create_uniquekey(uniqueKeyinfo: json,
                     sheet: object):

    if uniqueKeyinfo is None or len(uniqueKeyinfo)==0:
        return
    sheet.append(["唯一键信息"])
    for item in uniqueKeyinfo:
        content = []
        for field in item['fields']:
            content.append(field)
        sheet.append([",".join(content)])


def create_join_info(joinInfoList: json,
                     sheet: object):
    if joinInfoList is None or len(joinInfoList) == 0:
        return
    sheet.append(["外键引用"])
    sheet.append(["源字段", "目标库", "目标表", "目标字段"])
    for item in joinInfoList:
        content = []
        sourceFieldContent = []
        for field in item['fields']:
            sourceFieldContent.append(field)
        content.append(",".join(sourceFieldContent))
        content.append(item["refDb"])
        content.append(item["refTable"])
        tagetFieldContent = []
        for field in item["refFields"]:
            tagetFieldContent.append(field)
        content.append(",".join(tagetFieldContent))
        sheet.append(content)



def create_partition_info(partitionFieldInfoList: json,
                     sheet: object):
    if partitionFieldInfoList is None or len(partitionFieldInfoList) == 0 :
        return
    sheet.append(["分区信息"])
    sheet.append(["字段", "类型", "备注"])
    for item in partitionFieldInfoList:
        content = [item["name"], item["type"], item["remarks"]]
        sheet.append(content)

def update_schema():
    if len(sys.argv) < 3:
        print(
            "请输入excel路径："
        )
        filename = prompt().strip()
    else:
        filename = sys.argv[2]
    sq = sqldrafter.sqlDrafter()
    wb = load_workbook(filename=filename)
    if wb.sheetnames.count("空间") < 0:
        print("sheet名必须包含空间")
    spaceId = 0
    spaceName = ""
    busiDesc = ""
    tableItemDtoList = []
    for item in wb.sheetnames:
        sheet = wb[item]

        if item == "空间":
            for row in sheet.iter_rows(values_only=True):
                if "ID" in row:
                    spaceId = row[1]
                elif "标题：" in row:
                    spaceName = row[1]
                elif "描述：" in row:
                    busiDesc = row[1]
        else:
            tableItemDto = analysis_table(sheet)
            tableItemDtoList.append(tableItemDto)
    if spaceId.isdigit() and int(spaceId) > 0:
        spaceDto = SpaceEditReqDTO(spaceId, spaceName, busiDesc, tableItemDtoList)
        sq.edit_space(json.dumps(spaceDto, cls=JSONSelfEncoder))
    else:
        print("excel内容不合法，请用gen命令重新生成excel")
    sys.exit(0)



def analysis_table(sheet: object,
                    ):
    tableID = 0
    tableCnName = ""
    dbname=""
    fields = []
    uniq = []
    joinInfo = []
    partion = []
    listParam = []
    for row in sheet.iter_rows(values_only=True):
        if "ID：" in row:
            tableID = row[1]
        elif "表名" in row:
            name = row[1]
        elif "数据库名" in row :
            dbname =row[1]
        elif "中文名：" in row:
            tableCnName = row[1]
        elif "业务描述" in row:
            tableDesc = row[1]
        elif "表格详情" in row:
            analysis = analysisField
            listParam = fields
        elif "唯一键信息" in row:
            analysis = analysisUniq
            listParam = uniq
        elif "外键引用" in row:
            analysis = analysisJoinInfo
            listParam = joinInfo
        elif "分区信息" in row:
            analysis = analysisPartion
            listParam = partion
        else:
            analysis.analysis(row, listParam)

    tableItemDto = TableItemDto(tableID, 0, dbname, tableCnName, name, tableDesc, fields, joinInfo, uniq, partion)
    return tableItemDto

def update_indicator():
    if len(sys.argv) < 3:
        print("请输入空间ID：")
        space_id = prompt().strip()
    else:
        space_id = sys.argv[2]
    if space_id.isdigit():
        pass
    else:
        print("空间ID须为数字")
        sys.exit(0)
    if len(sys.argv)!=5:
        print("请输入正确的命令格式 sqldrafter indicator <spaceID> <key> <value>")
        sys.exit(0)
    sq = sqldrafter.sqlDrafter()
    detail = sq.get_detail(int(space_id))
    indicatorItemList= detail["data"]["indicatorItemList"]
    is_repalce=False;
    for item in indicatorItemList :
        if item["key"] == sys.argv[3]:
            item["desc"] =sys.argv[4]
            is_repalce =True
            break;
    if is_repalce:
        pass
    else:
        add_item={}
        add_item["key"]=sys.argv[3]
        add_item["desc"] =sys.argv[4]
        indicatorItemList.append(add_item)
    print(indicatorItemList)
    sq.edit_space(json.dumps({
        "id":space_id,
         "indicatorItemList":indicatorItemList
    }))
    sys.exit(0)
def to_str(field) -> str:
    if isinstance(field, str):
        return field
    elif isinstance(field, int):
        return str(field)
    elif isinstance(field, float):
        return str(field)
    elif isinstance(field, datetime.datetime):
        return field.strftime("%Y-%m-%d")
    elif isinstance(field, datetime.date):
        return field.strftime("%Y-%m-%d")
    elif isinstance(field, datetime.timedelta):
        return str(field)
    elif isinstance(field, datetime.time):
        return field.strftime("%H:%M:%S")
    elif isinstance(field, list):
        return str(field)
    elif isinstance(field, dict):
        return str(field)
    elif isinstance(field, bool):
        return str(field)
    elif isinstance(field, decimal.Decimal):
        return str(field)
    elif field is None:
        return "NULL"
    else:
        raise ValueError(f"Unknown type: {type(field)}")


def is_None(field) -> object:
    if field is None:
        return ''
    else:
        return field

def query():
    """
    初始化数据源配置在 ~/.sqldrafter/connection.json
    """
    if len(sys.argv) == 4:
        pass
    else:
        print(" 请输入正确格式 sqldrafter query <spaceID> <question>")
        sys.exit(0)
    sq = sqldrafter.sqlDrafter()
    if sys.argv[2].isdigit():
        pass
    else:
        print("空间ID必须为数字")
        sys.exit(0)
    spaceID = int(sys.argv[2])

    question =sys.argv[3]
    print(sq.query(spaceID,question))


    sys.exit(0)
class TableItemDto:
    def __init__(self, id, confirmOrNot, dbName, tableCnName, tableName, tableDesc, fields, joinInfoList,
                 uniqueKeyInfoList, partitionFieldInfoList):
        self.id = id
        self.confirmOrNot = confirmOrNot
        self.dbName = dbName
        self.tableCnName = tableCnName
        self.tableName = tableName
        self.tableDesc = tableDesc
        self.fields = fields
        self.joinInfoList = joinInfoList
        self.uniqueKeyInfoList = uniqueKeyInfoList
        self.partitionFieldInfoList = partitionFieldInfoList


class SpaceEditReqDTO:
    def __init__(self, id, name, busiDesc, tableList):
        self.id = id
        self.name = name
        self.busiDesc = busiDesc
        self.tableList = tableList


class JSONSelfEncoder(json.JSONEncoder):
    def default(self, obj):
        return obj.__dict__


if __name__ == "__main__":
    main()
