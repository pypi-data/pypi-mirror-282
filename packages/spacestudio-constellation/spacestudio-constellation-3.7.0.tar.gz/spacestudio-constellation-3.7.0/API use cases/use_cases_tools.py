from math import pi
from numpy import log as ln
import spacestudio
import openpyxl
from openpyxl import Workbook
import os
from pathlib import Path
import operator

# ---------------------------------------------------------------------------------------------------- object generation


def createSpacecraft(
    name,
    platformMass,
    platformPower,
    thrust,
    isp,
    thrusterPower,
    thrusterMass,
    propellantMass,
    solarArrayPower,
    nominalCapacity,
    standbyPower=0,
    warmUpDuration=0,
    warmUpPower=0,
    x=1,
    y=1,
    z=1,
    inertialCenterInSpacecraftFrameX=None,
    inertialCenterInSpacecraftFrameY=None,
    inertialCenterInSpacecraftFrameZ=None,
    solarArrayType="SOLAR_ARRAY_TYPE_DEPLOYABLE_FIXED",
    solarArrayEfficiency=0.3,
    depthOfDischarge=0.3,
    minimumChargeForFiring=0.9,
    powerSystemDefined="TRUE",
    dragModelDefined="FALSE",
    dragCoefficient=2.2,
    srpModelDefined="FALSE",
    attitudeModelDefined="FALSE",
    maxAngularVelocity=1 * (pi / 180),
    maxAngularAcceleration=5 * (pi / 180),
):

    thruster = {
        "name": name,
        "thrust": thrust,
        "isp": isp,
        "power": thrusterPower,
        "standbyPower": standbyPower,
        "warmUpDuration": warmUpDuration,
        "warmUpPower": warmUpPower,
        "propellantCapacityChoice": "PROPELLANT",
        "propellantMass": propellantMass,
        "totalMass": propellantMass + thrusterMass,
    }

    spacestudio.simulation.delete_propulsion_system(thruster["name"])
    spacestudio.simulation.create_propulsion_system(thruster)

    geometry = {
        "name": name,
        "type": "BOX",
        "x": x,
        "y": y,
        "z": z,
        "inertialCenterInSpacecraftFrameX": inertialCenterInSpacecraftFrameX,
        "inertialCenterInSpacecraftFrameY": inertialCenterInSpacecraftFrameY,
        "inertialCenterInSpacecraftFrameZ": inertialCenterInSpacecraftFrameZ,
        "thrusterAxisInSatelliteFrameX": 1.0,
        "thrusterAxisInSatelliteFrameY": 0.0,
        "thrusterAxisInSatelliteFrameZ": 0.0,
        "solarArrayType": solarArrayType,
        "solarArrayNormalInSatelliteFrameX": 0.0,
        "solarArrayNormalInSatelliteFrameY": 0.0,
        "solarArrayNormalInSatelliteFrameZ": -1.0,
        "solarArrayDefinitionType": "MAXIMUM_POWER",
        "solarArrayMaximumPower": solarArrayPower,
        "solarArrayEfficiency": solarArrayEfficiency,
    }

    spacestudio.simulation.delete_spacecraft_geometry(geometry["name"])
    spacestudio.simulation.create_spacecraft_geometry(geometry)

    battery = {
        "name": name,
        "nominalCapacity": nominalCapacity,
        "depthOfDischarge": depthOfDischarge,
        "minimumChargeForFiring": minimumChargeForFiring,
    }

    spacestudio.simulation.delete_battery(battery["name"])
    spacestudio.simulation.create_battery(battery)

    spacecraft = {
        "name": name,
        "platformMass": platformMass,
        "platformPower": platformPower,
        "onBoardAveragePower": thrusterPower,
        "dutyCycle": 0.7, 
        "powerSystemDefined": powerSystemDefined,
        "dragModelDefined": dragModelDefined,
        "dragCoefficient": dragCoefficient,
        "srpModelDefined": srpModelDefined,
        "attitudeModelDefined": attitudeModelDefined,
        "maxAngularVelocity": maxAngularVelocity,
        "maxAngularAcceleration": maxAngularAcceleration,
        "thruster": spacestudio.simulation.get_propulsion_system(thruster["name"]),
        "spacecraftGeometry": spacestudio.simulation.get_spacecraft_geometry(
            geometry["name"]
        ),
        "batteryDefinition": spacestudio.simulation.get_battery(battery["name"]),
    }

    spacestudio.simulation.delete_spacecraft(spacecraft["name"])
    spacestudio.simulation.create_spacecraft(spacecraft)


def createOrbit(
    name,
    type,
    sunSynchronousOrbit,
    altitude,
    ltan,
    inclination=None,
    perigeeAltitude=None,
    apogeeAltitude=None,
    anomaly=0,
    orbitDateTime="2022-01-01T12:00:00.000Z",
):

    initial_orbit = {
        "name": name,
        "type": type,
        "sunSynchronousOrbit": sunSynchronousOrbit,
        "perigeeAltitude": perigeeAltitude,
        "apogeeAltitude": apogeeAltitude,
        "altitude": altitude,
        "ascendingNodeDefinitionType": "LTAN",
        "ltan": ltan,
        "definitionType": "EllipticalWithPerigeeApogee",
        "inclination": inclination,
        "orbitDateTime": orbitDateTime,
        "meanOscType": "MEAN",
        "repeatOrbit": "FALSE",
        "perigeeArgument": 0,
        "anomalyType": "TRUE",
        "anomaly": anomaly,
    }

    spacestudio.simulation.delete_orbit(initial_orbit["name"])
    spacestudio.simulation.create_orbit(initial_orbit)


def createOrbitFromResults(results, newOrbitName):

    orbit = results["results"]["finalOrbit"]
    orbit["name"] = newOrbitName

    if orbit["type"] == "Elliptical":
        orbit["definitionType"] = "EllipticalWithPerigeeApogee"

    spacestudio.simulation.delete_orbit(orbit["name"])
    spacestudio.simulation.create_orbit(orbit)
    orbit = spacestudio.simulation.get_orbit(orbit["name"])
    return orbit


# ---------------------------------------------------------------------------------------------------- simulation mission generation


def createTransferMission(
    name,
    spacecraft,
    initialOrbit,
    targetDefinition,
    deltaSemiMajorAxis=None,
    deltaInclination=None,
    deltaEccentricity=0,
    targetOrbit=None,
    durationUntilPause=None,
    maximumDuration=600 * 86400,
    ephemeridesStepInSeconds=120,
    attitudeAndSystemSimulated=False,
    withEclipse=False,
    earthPotential=False,
    drag=False,
):

    mission_body = {
        "name": name,
        "spacecraft": spacecraft,
        "initialOrbit": initialOrbit,
        "targetDefinition": targetDefinition,
        "deltaSemiMajorAxis": deltaSemiMajorAxis,
        "deltaInclination": deltaInclination,
        "deltaEccentricity": deltaEccentricity,
        "targetOrbit": targetOrbit,
        "attitudeAndSystemSimulated": attitudeAndSystemSimulated,
        "withEclipse": withEclipse,
        "perturbations": [],
        "type": "Orbital Transfer",
        "propagationType": "NUMERICAL",
        "durationUntilPause": durationUntilPause,
        "maximumDuration": maximumDuration,
        "ephemeridesStepInSeconds": ephemeridesStepInSeconds,
    }

    if earthPotential == True:
        mission_body["perturbations"].append("EARTH_POTENTIAL")

    if drag == True:
        mission_body["perturbations"].append("DRAG")

    spacestudio.simulation.delete_mission(name)
    spacestudio.simulation.create_mission(mission_body)


def createRaanMission(
    name,
    spacecraft,
    initialOrbit,
    optimizationType,
    targetDeltaRaan,
    targetDefinition="DELTA",
    deltaSemiMajorAxis=0,
    deltaInclination=0,
    deltaEccentricity=0,
    targetOrbit=None,
    targetMissionDuration=None,
    targetDeltaVmax=None,
    durationUntilPause=None,
    maximumDuration=600 * 86400,
    ephemeridesStepInSeconds=86400 / 2,
    attitudeAndSystemSimulated=False,
    withEclipse=False,
    drag=False,
):
    mission_body = {
        "name": name,
        "spacecraft": spacecraft,
        "initialOrbit": initialOrbit,
        "targetDefinition": targetDefinition,
        "deltaSemiMajorAxis": deltaSemiMajorAxis,
        "deltaEccentricity": deltaEccentricity,
        "targetOrbit": targetOrbit,
        "deltaInclination": deltaInclination,
        "raanPhasingReference": "PHASING_WRT_DIRECT_TRANSFER",
        "optimizationType": optimizationType,
        "targetRaanLtanType": "RAAN",
        "targetDeltaRaan": targetDeltaRaan,
        "targetMissionDuration": targetMissionDuration,
        "targetDeltaVmax": targetDeltaVmax,
        "withEclipse": withEclipse,
        "attitudeAndSystemSimulated": attitudeAndSystemSimulated,
        "type": "RAAN Phasing",
        "propagationType": "Numerical",
        "durationUntilPause": durationUntilPause,
        "maximumDuration": maximumDuration,
        "ephemeridesStepInSeconds": ephemeridesStepInSeconds,
        "perturbations": [],
    }

    if drag == True:
        mission_body["perturbations"].append("DRAG")

    spacestudio.simulation.delete_mission(name)
    spacestudio.simulation.create_mission(mission_body)


def createStationKeepingMission(name,
    spacecraft,
    initialOrbit,
   smaTopMargin,
   smaBottomMargin,
   missionDuration,
   customSolarFlux = 150,
    durationUntilPause=None,
    maximumDuration=366 * 86400,
    ephemeridesStepInSeconds =120,
    attitudeAndSystemSimulated=False,
    withEclipse=False,
    earthPotential=True,
):

    mission_body = {
        "name": name,
        "smaTopMargin": smaTopMargin,
        "smaBottomMargin": smaBottomMargin,
        "duration": missionDuration,
        "spacecraft": spacecraft,
        "initialOrbit": initialOrbit,
        "toleranceType": "SMA_TOLERANCE",
        "attitudeAndSystemSimulated": attitudeAndSystemSimulated,
        "withEclipse": withEclipse,
        "perturbations": [],
        "type": "LEO Station Keeping",
        "propagationType": "Numerical",
        "durationUntilPause": durationUntilPause,
        "maximumDuration": maximumDuration,
        "ephemeridesStepInSeconds": ephemeridesStepInSeconds,
        "customSolarFlux": customSolarFlux,
        "propagationType": "Numerical",
    }
    mission_body["perturbations"].append("DRAG")
    if earthPotential == True:
        mission_body["perturbations"].append("EARTH_POTENTIAL")


    spacestudio.simulation.delete_mission(name)
    spacestudio.simulation.create_mission(mission_body)

# ---------------------------------------------------------------------------------------------------- mission computation


def computeAndGetResults(missionName):
    spacestudio.simulation.compute_mission(missionName)
    missionResults = spacestudio.simulation.wait_and_get_mission_results(missionName)
    return missionResults


# ---------------------------------------------------------------------------------------------------- functions relative to paused mission


def computeAndUpdateOAP(missionName, durationUntilPause):

    oaps = []

    body = spacestudio.simulation.get_mission(missionName)
    modified_body = body
    modified_body["durationUntilPause"] = durationUntilPause
    spacestudio.simulation.modify_mission(body, modified_body)

    spacecraft = spacestudio.simulation.get_spacecraft(body["spacecraft"]["name"])
    orbit = spacestudio.simulation.get_orbit(body["initialOrbit"]["name"])

    modifiedSpacecraft = spacecraft
    oap = computeOAPWithMSD(spacecraft, orbit)
    oaps.append(oap)
    modifiedSpacecraft["onBoardAveragePower"] = oap
    spacestudio.simulation.modify_spacecraft(spacecraft, modifiedSpacecraft)

    subMissionDuration = durationUntilPause
    i = 1
    missionToComputeName = missionName
    missionResults = computeAndGetResults(missionName=missionToComputeName)

    while isPaused(subMissionDuration, durationUntilPause) and i < 100:

        body = spacestudio.simulation.get_mission(missionToComputeName)

        missionResults = computeAndGetResults(missionName=missionToComputeName)
        subMissionDuration = missionResults["results"]["maneuverReport"][
            "simulationDuration"
        ]

        if isPaused(subMissionDuration, durationUntilPause):
            i += 1
            nextMissionName = missionName + " (pt." + str(i) + ")"
            spacestudio.simulation.delete_mission(nextMissionName)
            spacestudio.simulation.create_mission_from_paused_results(
                missionToComputeName, nextMissionName
            )
            missionToComputeName = nextMissionName

            currentMission = spacestudio.simulation.get_mission(missionToComputeName)

            spacecraftName = currentMission["spacecraft"]["name"]
            orbitName = currentMission["initialOrbit"]["name"]
            spacecraft = spacestudio.simulation.get_spacecraft(spacecraftName)
            orbit = spacestudio.simulation.get_orbit(orbitName)
            modifiedSpacecraft = spacecraft
            oap = computeOAPWithMSD(spacecraft, orbit)
            oaps.append(oap)
            modifiedSpacecraft["onBoardAveragePower"] = oap
            spacestudio.simulation.modify_spacecraft(spacecraft, modifiedSpacecraft)

    return oaps


def isPaused(subMissionDuration, durationUntilPause):
    durationTolerance = 1
    return (
        subMissionDuration <= durationUntilPause + durationTolerance
        and subMissionDuration >= durationUntilPause - durationTolerance
    )


# ---------------------------------------------------------------------------------------------------- functions relative to OAP orbital average power


def computeOAPWithMSD(spacecraft, orbit, msd_mission_name="tsp MSD"):
    msd_mission_body = {
        "name": msd_mission_name,
        "type": "Maneuvering Strategy Design",
        "spacecraft": spacecraft,
        "initialOrbit": orbit,
        "ephemeridesStepInSeconds": 60,
        "thrustAttitudeMode": "PROGRADE",
        "restAttitudeMode": "SUN_POINTING",
        "attitudeAndSystemSimulated": "True",
        "withEclipse": "True",
    }
    spacestudio.simulation.delete_mission(msd_mission_name)
    spacestudio.simulation.create_mission(msd_mission_body)
    spacestudio.simulation.compute_mission(msd_mission_name)
    rawResults = spacestudio.simulation.wait_and_get_mission_results(msd_mission_name)
    power_thrustMode = rawResults["results"]["thrustModeMeanPower"]
    power_restMode = rawResults["results"]["restModeMeanPower"]
    power_platform = spacecraft["platformPower"]
    power_thruster = spacecraft["thruster"]["power"]
    oap = (
        power_thruster
        * (power_restMode - power_platform)
        / (power_thruster + power_restMode - power_thrustMode)
    )
    print(spacecraft["name"], "OAP (W) :", oap, "\n")
    return oap


def updateSpacecraftOAP(spacecraft, orbit):
    oap = computeOAPWithMSD(spacecraft=spacecraft, orbit=orbit)
    dutyCycle = oap / spacecraft["thruster"]["power"]
    modifiedSpacecraft = spacecraft
    modifiedSpacecraft["onBoardAveragePower"] = oap
    modifiedSpacecraft["dutyCycle"] = dutyCycle
    spacestudio.simulation.modify_spacecraft(spacecraft, modifiedSpacecraft)
    return oap


# ---------------------------------------------------------------------------------------------------- other functions


def getMaxDeltaV(spacecraft):
    thruster = spacecraft["thruster"]
    propellantMass = thruster["propellantMass"]
    thrusterMass = thruster["totalMass"] - thruster["propellantMass"]
    platformMass = spacecraft["platformMass"]
    dryMass = platformMass + thrusterMass
    wetMass = dryMass + propellantMass
    isp = thruster["isp"]
    g0 = 9.80665
    maxDeltaV = isp * g0 * ln((wetMass / dryMass))
    print("Maximum ΔV :", round(maxDeltaV, 2), "m/s", "\n")
    return maxDeltaV


def modifyPropellantMass(spacecraft, mass):
    spacecraft = spacestudio.simulation.get_spacecraft(spacecraft["name"])
    thruster = spacecraft["thruster"]
    new_thruster = thruster
    new_thruster["propellantMass"] = thruster["propellantMass"] - mass
    new_thruster["totalMass"] = thruster["totalMass"] - mass
    spacestudio.simulation.modify_propulsion_system(thruster, new_thruster)
    print(spacecraft["name"], "propellant mass updated", "\n")
    pass


def export_mission_results_in_xlsx(mission):

  result = spacestudio.simulation.get_mission_results(mission)
  
  name = result['results']['name'] + '_results.xlsx'
  dirname = str(os.path.join(Path.home(), "Downloads"))
  xlsxFile = os.path.join(dirname, name)

  print('writing', name, '  ...')

  wb = Workbook()
  ws1 = wb.active
  ws1.title = "MissionId"
  ws1.column_dimensions[openpyxl.utils.cell.get_column_letter(1)].width = 40
  ws1.column_dimensions[openpyxl.utils.cell.get_column_letter(2)].width = 50
  sheet = 'MissionId'

  dico = result
  lst_dico = []
  lst_dico.append(dico)

  def iterxlsxx(d):
    nonlocal xlsxFile
    nonlocal sheet
    nonlocal lst_dico
    nonlocal wb

    for k,v in d.items():
      if isinstance(v, dict):

        if k in list(lst_dico[-1].keys()):
          wb.active = wb[sheet]
          ws = wb.active
          ws.append([k, 'See '+k+' sheet'])
        else:
          i = len(lst_dico)
          while k not in list(lst_dico[i-1].keys()):
            i = i-1
          else:
            sheetname = wb.sheetnames[i-1]
            wb.active = wb[sheetname]
            ws = wb.active
            ws.append([k, 'See '+k+' sheet'])

        ws = wb.create_sheet(title=k)
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(1)].width = 40
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(2)].width = 50

        sheet = k
        lst_dico.append(v)
        iterxlsxx(v)
      
      else:
        if k in list(lst_dico[-1].keys()):
          wb.active = wb[sheet]
          ws = wb.active
          if type(v) == list:
            v = str(v)
          ws.append([k, v])
        
        else:
          i = len(lst_dico)
          while k not in list(lst_dico[i-1].keys()):
            i = i-1
          else:
            sheetname = wb.sheetnames[i-1]
            wb.active = wb[sheetname]
            ws = wb.active
            if type(v) == list:
              v = str(v)
            ws.append([k, v]) 
    
  iterxlsxx(result)

  if(result['results']['fieldIndexes'] != None):
    ws = wb.create_sheet('ephemerides')

    headers = sorted(result['results']['fieldIndexes'].items(), key=operator.itemgetter(1))
    headers = dict(headers)
    headers = list(headers.keys())
    ws.append(headers)

    lst = result['results']['ephemerides']
    for i in range(len(lst)):
      ws.append(lst[i])
    
    for i in range(len(headers)):
      ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i+1)].width = 20
  
  if len(result['errors']) != 0:
    wb.active = wb['MissionId']
    ws = wb.active

    for i in range(len(result['errors'])):
        message = result['errors'][i]['message']
        ws.append(['Error message n° '+str(i), message])


  wb.save(filename = xlsxFile)
  print(name, 'successfully created at',dirname)


def getEphemerides(missionName, ephemerideType):
    missionResults = spacestudio.simulation.get_mission_results(missionName)
    ephemerides = []
    if all(
        x in missionResults["results"].keys() for x in ["ephemerides", "fieldIndexes"]
    ):
        if ephemerideType in missionResults["results"]["fieldIndexes"].keys():

            ephemerideNumber = missionResults["results"]["fieldIndexes"][ephemerideType]
            for ephemeride in missionResults["results"]["ephemerides"]:
                ephemerides.append(ephemeride[ephemerideNumber])

        else:
            print("ERROR: can't find", ephemerideType, "in the fieldIndexes")
    else:
        print(
            "ERROR: can't find ephemerides or fieldIndexes arguments in missionResults"
        )
    return ephemerides


# ----------------------------------------------------------------------------------------------------
